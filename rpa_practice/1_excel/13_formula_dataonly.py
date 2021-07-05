# 수식2 (데이터만 가져오기)

from openpyxl import load_workbook

# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# 수식 그대로 가져옴 (수식 값이 아니라)
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가지고 옴.
# evaluate 되지 않은 상태의 데이터는 None이라고 표시 -> 파일을 열어서 저장했다가 다시 실행하면 정상적으로 표시됨.
for row in ws.values:
    for cell in row:
        print(cell)
