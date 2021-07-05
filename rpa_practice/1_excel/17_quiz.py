from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 미리 작성된 데이터 넣기
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))
scores = [
    (1, 10, 8, 5, 14, 26, 12),
    (2, 7, 3, 7, 15, 24, 18),
    (3, 9, 5, 8, 8, 12, 4),
    (5, 7, 8, 7, 17, 21, 18),
    (6, 3, 5, 8, 8, 17, 0),
    (7, 4, 9, 10, 16, 27, 18),
    (8, 6, 6, 6, 15, 19, 17),
    (9, 10, 10, 9, 19, 30, 19),
    (10, 9, 8, 8, 20, 25, 20)
]
for score in scores:
    ws.append(score)

# 퀴즈2 점수 전부 10으로 수정
for row in range(2, ws.max_row+1):
    ws.cell(row, 4, value=10)

# H 열에 총점, I 열에 성적 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):
    sum_score = sum(score[1:]) - score[3] + 10  # 총점
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)

    grade = None                                # 성적
    if sum_score >= 90:
        grade = "A"
    elif sum_score >= 80:
        grade = "B"
    elif sum_score >= 70:
        grade = "C"
    else:
        grade = "D"

    if ws.cell(row=idx, column=2).value < 5:
        grade = "F"

    ws.cell(row=idx, column=9).value = grade

wb.save("scores.xlsx")